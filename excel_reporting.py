from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.formatting.rule import CellIsRule
import os.path

class WorkbookManager:
	def __init__(self, workbook, percentage_formatting):
		self.workbook = workbook
		self.percentage_formatting = percentage_formatting

	def write_results_to_worksheet(self, test_results, sheet_name, is_new_sheet=False, table_name=None,
		table_title=None, is_failures_reported=True, is_main_sheet=False):
		worksheet = self.workbook.active
		if is_new_sheet:
			wb = self.workbook
			if is_main_sheet:
				worksheet = wb.create_sheet(index=0)
			else:
				worksheet = wb.create_sheet()
		worksheet.title = sheet_name

		starting_row = ResultsTable.STARTING_ROW
		if table_title:
			excel_mgr = WorksheetManager(worksheet)
			excel_mgr.paint_cell(starting_row, FailureList.COLUMN, table_title, is_bold=True, font_size=14)
			starting_row += 2

		results_table_name = (table_name or sheet_name)
		table = ResultsTable(results_table_name, self.percentage_formatting, starting_row=starting_row)
		for result in test_results:
			table.add_result(result['app_title'], result['number_passing'], result['number_failing'])
		next_row = table.write_results(worksheet, is_main_sheet) + 1

		if is_failures_reported:
			self._write_failures_to_worksheet(worksheet, next_row, test_results)

	def save_workbook(self):
		default_filename = 'regression_run'
		filename = input('Enter filename (\'' + default_filename + '\'): ') or default_filename
		extension = '.xlsx'
		filename = filename.replace(extension, '') + extension

		is_saving = True
		if os.path.isfile(filename):
			user_input = input('\'' + filename + '\' already exists. Would you like to overwrite it (Y/N)? ')
			is_saving = (user_input in ['Y', 'y', 'Yes', 'YES', 'yes'])
		if is_saving:
			print('Saving \'' + filename + '\'...')
			self.workbook.save(filename=(filename))
			print('Saved')

	def _write_failures_to_worksheet(self, worksheet, next_row, test_results):
		excel_mgr = WorksheetManager(worksheet)
		excel_mgr.paint_cell(next_row, FailureList.COLUMN, 'Failures', is_bold=True, is_underline=True, font_size=14)
		next_row += 1

		for result in test_results:
				failures = FailureList(result['app_title'], result['failure_links'], self.percentage_formatting)
				next_row = failures.write_results(worksheet, next_row + 1)

class WorksheetManager:
	def __init__(self, worksheet):
		self.worksheet = worksheet

	@staticmethod
	def fill_from_hex_value(hex_value):
		return PatternFill(start_color=hex_value, end_color=hex_value, fill_type='solid')

	def paint_cell(self, row, col, text, fill_color=None, is_bold=False, is_italic=False, is_underline=False, 
		font_size=12, border=None, is_percent=False, font_color='000000'):
		cell_location = chr(col) + str(row)
		self.worksheet[cell_location] = text
		if fill_color:
			self.worksheet[cell_location].fill = WorksheetManager.fill_from_hex_value(fill_color)
		if border:
			self.worksheet[cell_location].border = border
		if is_percent:
			self.worksheet[cell_location].number_format = '0%'
		underline_value = 'single' if is_underline else None
		self.worksheet[cell_location].font = Font(bold=is_bold, italic=is_italic, underline=underline_value, size=font_size, color=font_color)

	def paint_hyperlink(self, row, col, link, border=None, font_size=12):
		cell_location = chr(col) + str(row)
		self.worksheet[cell_location] = link['value']
		self.worksheet[cell_location].hyperlink = link['url']
		self.worksheet[cell_location].font = Font(underline='single', color='0563C1', size=font_size)
		if border:
			self.worksheet[cell_location].border = border

class ResultsTable:
	STARTING_ROW = 2
	STARTING_COL = 66
	BORDER = Border(left=Side(style='thin',color='000000'),
		right=Side(style='thin',color='000000'),
		top=Side(style='thin',color='000000'),
		bottom=Side(style='thin',color='000000'))

	def __init__(self, job_name, percentage_formatting, starting_row=None):
		self.job_name = job_name
		self.percentage_formatting = percentage_formatting
		self.results = []
		self.starting_row = (starting_row or ResultsTable.STARTING_ROW)

	def add_status_formatting_to_range(self, ws, format_range):
		for format_type, frmt in self.percentage_formatting.items():
			ws.conditional_formatting.add(format_range, CellIsRule(operator=frmt['range']['operator'], 
				formula=frmt['range']['value'], fill=WorksheetManager.fill_from_hex_value(frmt['fill_color']),
				font=Font(color=frmt['font_color'])))

	def add_result(self, app_name, passCount, failCount):
		self.results.append({'app_name': app_name, 'passCount': passCount, 'failCount': failCount})

	def write_results(self, worksheet, is_apps_linked=False):
		headerFill = 'A9D08E'
		totalFill = 'E7E6E6'

		percent_col = chr(self.STARTING_COL + 4)
		avg_col = chr(ResultsTable.STARTING_COL + 5)
		ending_row = str(self.starting_row + len(self.results) + 1)
		format_range = percent_col + str(self.starting_row + 1) + ':' + percent_col + ending_row 
		self.add_status_formatting_to_range(worksheet, format_range)
		excel_mgr = WorksheetManager(worksheet)

		row = self.starting_row
		col = ResultsTable.STARTING_COL
		worksheet.column_dimensions[chr(col)].width = 35
		excel_mgr.paint_cell(row, col, self.job_name, fill_color=headerFill, font_size=14, border=ResultsTable.BORDER)
		col += 1
		excel_mgr.paint_cell(row, col, 'Total', fill_color=headerFill, font_size=14, border=ResultsTable.BORDER)
		col += 1
		excel_mgr.paint_cell(row, col, 'Passing', fill_color=headerFill, font_size=14, border=ResultsTable.BORDER)
		col += 1
		excel_mgr.paint_cell(row, col, 'Failing', fill_color=headerFill, font_size=14, border=ResultsTable.BORDER)
		col += 1
		worksheet.column_dimensions[chr(col)].width = 20
		excel_mgr.paint_cell(row, col, 'Percent Passing', fill_color=headerFill, font_size=14, border=ResultsTable.BORDER)

		col = ResultsTable.STARTING_COL
		row += 1
		starting_row = row
		for result in self.results:
			total_number_tests = float(result['passCount']) + float(result['failCount'])
			self._paint_app_name(excel_mgr, row, col, is_apps_linked, result['app_name'])
			col += 1
			excel_mgr.paint_cell(row, col, (total_number_tests), border=ResultsTable.BORDER)
			col += 1
			excel_mgr.paint_cell(row, col, result['passCount'], border=ResultsTable.BORDER)
			col += 1
			excel_mgr.paint_cell(row, col, result['failCount'], border=ResultsTable.BORDER)
			col += 1
			excel_mgr.paint_cell(row, col, (result['passCount'] / total_number_tests), border=ResultsTable.BORDER, is_percent=True)
			row += 1
			col = ResultsTable.STARTING_COL

		totalFill = 'E7E6E6'
		excel_mgr.paint_cell(row, col, 'TOTAL', fill_color=totalFill, is_bold=True, border=ResultsTable.BORDER)
		for i in range(0,3):
			col += 1
			cell_range = chr(col) + str(starting_row) + ':' + chr(col) + str(row-1)
			excel_mgr.paint_cell(row, col, '=SUM(' + cell_range + ')', fill_color=totalFill, is_bold=True, border=ResultsTable.BORDER)
		percentage_passed_formula = '=' + chr(col-1) + str(row) + '/' + chr(col-2) + str(row)
		excel_mgr.paint_cell(row, col+1, percentage_passed_formula, border=ResultsTable.BORDER, is_percent=True)

		formula_range = chr(col+1) + str(starting_row) + ':' + chr(col+1) + str(row-1)
		avg_passed_formula = '=AVERAGE(' + formula_range + ')'
		stdev_passed_formula = '=STDEV.P(' + formula_range + ')'
		excel_mgr.paint_cell(row-1, col+2, 'Avg. % Pass', fill_color='000000', font_color='FFFFFF', is_bold=True)
		excel_mgr.paint_cell(row, col+2, avg_passed_formula, border=ResultsTable.BORDER, is_percent=True)
		self.add_status_formatting_to_range(worksheet, (chr(col+2) + str(row) + ':' + chr(col+2) +str(row)))
		worksheet.column_dimensions[chr(col+2)].width = 15

		return row + 1

	def _paint_app_name(self, excel_mgr, row, col, is_apps_linked, app_name):
		if is_apps_linked:
			link = { 'value': app_name, 'url': '#\'' + app_name + '\'!B2' }
			excel_mgr.paint_hyperlink(row, col, link, border=ResultsTable.BORDER)
		else:
			excel_mgr.paint_cell(row, col, app_name, font_size=14, border=ResultsTable.BORDER)

class FailureList:
	COLUMN = ResultsTable.STARTING_COL

	def __init__(self, app_name, failures, percentage_formatting):
		self.app_name = app_name
		self.failures = failures
		self.percentage_formatting = percentage_formatting

	def write_results(self, worksheet, row_number):
		excel_mgr = WorksheetManager(worksheet)
		excel_mgr.paint_cell(row_number, FailureList.COLUMN, self.app_name)
		row = row_number + 1
		if len(self.failures) == 0:
			success_font = self.percentage_formatting['success']['font_color']
			success_fill = self.percentage_formatting['success']['fill_color']
			excel_mgr.paint_cell(row, FailureList.COLUMN, 'No Failures', is_italic=True, fill_color=success_fill,
				font_color=success_font)
			row += 1
		else:
			for link in self.failures:
				excel_mgr.paint_hyperlink(row, FailureList.COLUMN, link)
				row += 1
		return row